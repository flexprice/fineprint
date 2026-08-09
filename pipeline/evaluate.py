"""Score a pipeline result against the hand-filled ground-truth breakdown row.

  python3 -m pipeline.evaluate "<contract name or folder>" [result.json]

Loads the matching row from the ground-truth workbook, maps it to our contract-value
fields, and prints a per-field match table + accuracy. Used to drive prompt/logic
refinement. Point it at your own labeled workbook via ``FINEPRINT_GROUND_TRUTH``.
"""
import json
import os
import re
import sys
from pathlib import Path

import openpyxl

from .config import PROJECT_ROOT

XLSX = Path(os.environ.get("FINEPRINT_GROUND_TRUTH", PROJECT_ROOT / "data" / "ground_truth.xlsx"))

# breakdown column (1-based) -> our field name
COL = {
    6: "start_date",
    8: "platform_fee.amount", 9: "platform_fee.frequency", 10: "platform_fee.timing",
    12: "hosting_fee.amount", 13: "hosting_fee.frequency", 14: "hosting_fee.timing",
    16: "llm_usage_fee.amount", 17: "llm_usage_fee.frequency", 18: "llm_usage_fee.timing",
    20: "credit_grant.amount", 21: "credit_grant.type",
    22: "entitlement.description", 23: "entitlement.period",
    24: "override_hosting_per_min", 25: "override_sms_per_msg", 26: "override_other",
    27: "commitment.amount", 28: "commitment.period", 29: "commitment.overage_factor",
    30: "commitment.true_up", 31: "commitment.scope_notes",
    32: "payment_terms",
}
NUMERIC = {"platform_fee.amount", "hosting_fee.amount", "llm_usage_fee.amount",
           "credit_grant.amount", "override_hosting_per_min", "override_sms_per_msg",
           "commitment.amount", "commitment.overage_factor"}
# free-text fields: reviewed by a human, not scored by exact string match
SOFT = {"entitlement.description", "override_other", "commitment.scope_notes", "payment_terms"}


def _canon_date(s):
    s = str(s).strip()
    for fmt in ("%Y-%m-%d", "%d%b%Y", "%d %b %Y", "%b %d, %Y", "%B %d, %Y",
                "%m/%d/%Y", "%d-%b-%Y", "%d %B %Y", "%Y%m%d"):
        try:
            from datetime import datetime
            return datetime.strptime(s[:len(fmt) + 4], fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    m = re.search(r"(\d{4})-(\d{2})-(\d{2})", s)
    return m.group(0) if m else s.lower()


def _norm(field, v):
    s = "" if v is None else str(v).strip()
    empty = s.lower() in ("", "n/a", "none", "nan")
    if field in NUMERIC:                       # 0 and missing are equivalent (no such charge)
        if empty:
            return 0.0
        m = re.search(r"[-+]?\d[\d,]*\.?\d*", s)   # extract number from "$0.05/min" etc.
        if m:
            return round(float(m.group(0).replace(",", "")), 6)
        return s.lower()
    if field == "commitment.true_up":          # absent true-up defaults to "no"
        return "yes" if s.lower() in ("yes", "true") else "no"
    if empty:
        return None
    if field == "start_date":
        return _canon_date(s)
    return s.lower()


_ANNUAL_MULT = {"monthly": 12, "quarterly": 4, "half yearly": 2, "annual": 1}


def _annualized(row, group):
    amt = row.get(f"{group}.amount")
    if not amt:
        return None
    fr = row.get(f"{group}.frequency")
    return round(float(amt) * _ANNUAL_MULT.get(fr, 1), 2)


def load_truth(name: str):
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    ws = wb["Contract Breakdown"]
    key = name.lower().replace("(renewal)", "").replace("systems", "").strip()
    best = None
    for r in range(2, ws.max_row + 1):
        cid = ws.cell(r, 1).value
        if not cid:
            continue
        cid_l = str(cid).lower()
        if key[:6] in cid_l or cid_l[:6] in key:
            best = r
            break
    if not best:
        return None, None
    row = {}
    for c, f in COL.items():
        row[f] = _norm(f, ws.cell(best, c).value)
    return str(ws.cell(best, 1).value), row


def evaluate(name: str, result_path: str = None):
    truth_name, truth = load_truth(name)
    if not truth:
        print(f"No ground-truth row found for '{name}'")
        return
    rp = Path(result_path) if result_path else PROJECT_ROOT / "outputs" / name / "result.json"
    if not rp.exists():
        print(f"No result.json at {rp} (run the pipeline first)")
        return
    res = json.loads(rp.read_text())
    pred = {f["field"]: _norm(f["field"], f["value"]) for f in res["fields"]}

    print(f"\nGround truth row: {truth_name}\n")
    print(f"{'field':<32} {'expected':<24} {'predicted':<24} match")
    print("-" * 88)
    scored = correct = 0
    for f in COL.values():
        if f in SOFT:
            continue
        exp, got = truth.get(f), pred.get(f)
        if exp in (None, 0.0) and got in (None, 0.0):
            continue
        scored += 1
        ok = exp == got
        if not ok and f.rsplit(".", 1)[-1] in ("amount", "frequency"):
            g = f.rsplit(".", 1)[0]
            if g in ("platform_fee", "llm_usage_fee", "hosting_fee"):
                at, ap = _annualized(truth, g), _annualized(pred, g)
                if at is not None and at == ap:   # $10k/qtr == $40k/yr (economic equivalence)
                    ok = True
        correct += ok
        print(f"{f:<32} {str(exp)[:23]:<24} {str(got)[:23]:<24} {'✓' if ok else '✗'}")
    print("-" * 88)
    print(f"STRUCTURED accuracy: {correct}/{scored} = {correct/scored*100:.0f}%" if scored else "no comparable fields")

    print("\nfree-text fields (human review, not strict-scored):")
    for f in COL.values():
        if f in SOFT and (truth.get(f) or pred.get(f)):
            print(f"  {f}\n    expected : {truth.get(f)}\n    predicted: {pred.get(f)}")


if __name__ == "__main__":
    evaluate(sys.argv[1], sys.argv[2] if len(sys.argv) > 2 else None)
