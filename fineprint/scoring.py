"""Scoring — economic-substance field scoring for the contract→billing extraction task.

Two headline sub-metrics, reported separately (see methodology):
  • EXTRACTION — the economic facts a model must read off the page: start date, the *set* of
    fees (matched by annualized value so a platform/LLM bucket swap isn't punished six times),
    credit/commitment amounts, and per-unit overrides.
  • CONVENTION — house-rule fields whose value is largely a labeling convention (fee timing
    advanced/arrear, entitlement/commitment period, credit-grant type). Scored, but kept out of
    the extraction number so a model isn't penalized for a defensible reading of a convention.

Beyond accuracy we surface hallucination (a field marked HIGH-confidence but wrong on an
extraction fact). Ground truth is private; only aggregates leave here. The workbook labels can
be layered with cleaned GOLD_LABELS (the ensemble+adjudication label-QA pass).
"""
import json
import openpyxl

from pipeline import evaluate as ev
from fineprint.config import GROUND_TRUTH, GOLD_LABELS

_TRUTH: dict[str, dict] = {}
_GOLD: dict[str, dict] | None = None

# ── enum canonicalization (applied to BOTH truth and prediction) ────────────
_FREQ = {"annual": "annual", "annually": "annual", "yearly": "annual", "per annum": "annual",
         "per year": "annual", "12-month": "annual", "quarterly": "quarterly", "per quarter": "quarterly",
         "quarter": "quarterly", "per 90 days": "quarterly", "monthly": "monthly", "per month": "monthly",
         "half yearly": "half yearly", "semi-annual": "half yearly", "semiannual": "half yearly",
         "biannual": "half yearly", "bi-annual": "half yearly"}
_TIMING = {"advanced": "advanced", "advance": "advanced", "in advance": "advanced", "upfront": "advanced",
           "prepaid": "advanced", "arrear": "arrear", "arrears": "arrear", "in arrears": "arrear"}
_CREDIT = {"one time": "credit", "one-time": "credit", "onetime": "credit", "usage bank": "credit",
           "prepaid": "credit", "usage credit": "credit", "committed spend": "credit", "bank": "credit"}
_MULT = {"monthly": 12, "quarterly": 4, "half yearly": 2, "annual": 1}

GROUPS = ("recurring_fee", "fixed_fee", "usage_fee")
# field partitions
EXTRACT_SCALAR = ("start_date", "credit_grant.amount", "commitment.amount", "commitment.overage_factor",
                  "override_hosting_per_min", "override_sms_per_msg")
CONVENTION = ("recurring_fee.timing", "fixed_fee.timing", "usage_fee.timing",
              "entitlement.period", "commitment.period", "credit_grant.type", "commitment.true_up")


def _canon(kind, v):
    if v is None:
        return None
    s = str(v).strip().lower()
    if s in ("", "n/a", "none", "nan"):
        return None
    if kind == "freq":
        return _FREQ.get(s, s)
    if kind == "timing":
        return _TIMING.get(s, s)
    if kind == "credit":
        return _CREDIT.get(s, "credit")   # any stated grant type -> "credit present"
    return s


def _annualized(amount, freq):
    try:
        a = float(amount)
    except (TypeError, ValueError):   # model may emit a non-numeric amount — treat as absent
        return None
    if not a:
        return 0.0
    return round(a * _MULT.get(_canon("freq", freq), 1), 2)


def _fee_bag(row: dict) -> list[tuple]:
    """Multiset of (annualized value, canon timing) for each nonzero fee group — bucket-agnostic."""
    bag = []
    for g in GROUPS:
        av = _annualized(row.get(f"{g}.amount"), row.get(f"{g}.frequency"))
        if av:
            bag.append((av, _canon("timing", row.get(f"{g}.timing"))))
    return bag


def _match_bags(truth_bag, pred_bag, tol=0.02):
    """Greedy match by annualized value within tol. -> (matched, timing_ok, timing_total)."""
    used = [False] * len(pred_bag)
    matched = tim_ok = tim_tot = 0
    for tv, tt in truth_bag:
        for j, (pv, pt) in enumerate(pred_bag):
            if used[j] or tv is None or pv is None:
                continue
            if abs(pv - tv) <= max(1.0, abs(tv) * tol):
                used[j] = True
                matched += 1
                if tt is not None:
                    tim_tot += 1
                    tim_ok += (pt == tt)
                break
    return matched, tim_ok, tim_tot


def _load_gold() -> dict:
    global _GOLD
    if _GOLD is None:
        _GOLD = json.loads(GOLD_LABELS.read_text()) if GOLD_LABELS and GOLD_LABELS.exists() else {}
    return _GOLD


def load_truth(name: str) -> dict | None:
    """Exact-name ground-truth row -> {field: normalized_value}. Cached; loaded once.

    If cleaned GOLD_LABELS exist, its per-field corrections are layered over the workbook row.
    """
    if not _TRUTH:
        ws = openpyxl.load_workbook(GROUND_TRUTH, data_only=True)["Contract Breakdown"]
        for r in range(2, ws.max_row + 1):
            cid = ws.cell(r, 1).value
            if cid:
                _TRUTH[str(cid).strip()] = {f: ev._norm(f, ws.cell(r, c).value) for c, f in ev.COL.items()}
    row = _TRUTH.get(name)
    if row is None:
        return None
    gold = _load_gold().get(name)
    if gold:
        row = {**row, **{f: ev._norm(f, v) for f, v in gold.items()}}
    return row


def _score_core(pred: dict, conf: dict, truth: dict) -> dict:
    ex_c = ex_s = cv_c = cv_s = high = confident_wrong = 0
    # fees as a bag -> extraction; their timings -> convention
    tb, pb = _fee_bag(truth), _fee_bag(pred)
    matched, tim_ok, tim_tot = _match_bags(tb, pb)
    fee_scored = max(len(tb), len(pb))
    if fee_scored:
        ex_s += fee_scored
        ex_c += matched
    cv_s += tim_tot
    cv_c += tim_ok
    # scalar extraction facts
    for f in EXTRACT_SCALAR:
        exp, got = truth.get(f), pred.get(f)
        if exp in (None, 0.0) and got in (None, 0.0):
            continue
        ex_s += 1
        ok = exp == got
        ex_c += ok
        if conf.get(f) == "HIGH":
            high += 1
            confident_wrong += not ok
    # convention fields (non-timing; timings handled with the fee bag)
    for f in CONVENTION:
        if f.endswith(".timing"):
            continue
        kind = "credit" if f == "credit_grant.type" else "freq" if f.endswith("period") else "text"
        e = _canon(kind, truth.get(f))
        g = _canon(kind, pred.get(f))
        if e is None and g is None:
            continue
        cv_s += 1
        cv_c += (e == g)
    return {"ex_correct": ex_c, "ex_scored": ex_s, "cv_correct": cv_c, "cv_scored": cv_s,
            "high": high, "confident_wrong": confident_wrong,
            "correct": ex_c + cv_c, "scored": ex_s + cv_s}   # combined (aggregate's headline accuracy)


def score(fields, truth) -> dict:
    """Per-run scorecard: extraction + convention (correct/scored), combined, and HIGH-confidence
    hallucinations (confident and wrong on an extraction fact)."""
    pred = {f.field: ev._norm(f.field, f.value) for f in fields}
    conf = {f.field: f.confidence for f in fields}
    return _score_core(pred, conf, truth)


def score_detail(fields, truth) -> list[dict]:
    """Per-field audit rows (expected · predicted · raw · confidence · bucket · correct). Private."""
    pred = {f.field: ev._norm(f.field, f.value) for f in fields}
    raw = {f.field: f.value for f in fields}
    conf = {f.field: f.confidence for f in fields}
    rows = []
    # fee bag summary row
    tb, pb = _fee_bag(truth), _fee_bag(pred)
    matched, tim_ok, tim_tot = _match_bags(tb, pb)
    rows.append({"field": "fees(bag)", "expected": f"{len(tb)} fees", "predicted": f"{matched} matched/{len(pb)}",
                 "raw": None, "confidence": None, "bucket": "extraction",
                 "scored": bool(max(len(tb), len(pb))), "correct": matched == max(len(tb), len(pb))})
    for f in EXTRACT_SCALAR:
        exp, got = truth.get(f), pred.get(f)
        scored = not (exp in (None, 0.0) and got in (None, 0.0))
        rows.append({"field": f, "expected": exp, "predicted": got, "raw": raw.get(f),
                     "confidence": conf.get(f), "bucket": "extraction",
                     "scored": scored, "correct": (exp == got) if scored else None})
    for f in CONVENTION:
        kind = "credit" if f == "credit_grant.type" else "freq" if f.endswith("period") else "timing" if f.endswith(".timing") else "text"
        e, g = _canon(kind, truth.get(f)), _canon(kind, pred.get(f))
        scored = not (e is None and g is None)
        rows.append({"field": f, "expected": e, "predicted": g, "raw": raw.get(f),
                     "confidence": conf.get(f), "bucket": "convention",
                     "scored": scored, "correct": (e == g) if scored else None})
    return rows
